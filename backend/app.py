"""DIPROMES — Flask backend."""
import json
import os
from datetime import timedelta
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory, Response, session
from flask_cors import CORS
from collections import defaultdict
from time import time
from werkzeug.security import generate_password_hash, check_password_hash

from models import db, Maquina, Registro, Usuario, PacienteMaster, Config
from exportar import generate_excel, generate_csv, generate_pdf

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IS_PROD = bool(os.environ.get("DATABASE_URL"))

app = Flask(__name__, static_folder=BASE_DIR)

# ─── Security config ─────────────────────────────────────────────────────────
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", os.urandom(32).hex())
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = IS_PROD
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

# ─── CORS — only allow the production origin (same-origin in dev) ────────────
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "https://dipromes.onrender.com")
CORS(app, origins=[ALLOWED_ORIGIN] if IS_PROD else ["*"], supports_credentials=True)

# ─── HTTP security headers (no external lib needed) ──────────────────────────
@app.after_request
def add_security_headers(resp):
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    if IS_PROD:
        resp.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return resp

# ─── Rate limiting — simple in-memory token bucket per IP ────────────────────
# ponytail: replace with Redis-backed solution when scaling beyond 1 dyno
_login_attempts: dict = defaultdict(list)
_RATE_WINDOW = 60   # seconds
_RATE_MAX    = 10   # attempts per window

def _check_rate_limit(ip: str) -> bool:
    now = time()
    attempts = [t for t in _login_attempts[ip] if now - t < _RATE_WINDOW]
    _login_attempts[ip] = attempts
    if len(attempts) >= _RATE_MAX:
        return False
    _login_attempts[ip].append(now)
    return True

# ─── Database ────────────────────────────────────────────────────────────────
_db_url = os.environ.get(
    "DATABASE_URL",
    f"sqlite:///{os.path.join(BASE_DIR, 'meditrack.db')}"
)
if _db_url.startswith("postgres://"):
    _db_url = _db_url.replace("postgres://", "postgresql+pg8000://", 1)
elif _db_url.startswith("postgresql://") and "pg8000" not in _db_url:
    _db_url = _db_url.replace("postgresql://", "postgresql+pg8000://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = _db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)


# ─── Auth decorators ─────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return jsonify({"error": "No autorizado"}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return jsonify({"error": "No autorizado"}), 401
        if session.get("rol") != "admin":
            return jsonify({"error": "Se requiere rol de administrador"}), 403
        return f(*args, **kwargs)
    return decorated


# ─── ID helpers ──────────────────────────────────────────────────────────────

def next_num(model, prefix):
    rows = db.session.query(model.id).all()
    nums = [int(r[0][len(prefix):]) for r in rows
            if r[0].startswith(prefix) and r[0][len(prefix):].isdigit()]
    return max(nums, default=0) + 1


def next_reg_id():
    return f"R{next_num(Registro, 'R'):03d}"

def next_maq_id():
    return f"MAQ{next_num(Maquina, 'MAQ'):02d}"

def next_usr_id():
    return f"U{next_num(Usuario, 'U'):03d}"


# ─── Computed field ──────────────────────────────────────────────────────────

def paciente_actual(maq_id):
    r = Registro.query.filter_by(maquina=maq_id, estatus="Activo").first()
    return r.nombre if r else None


# ─── Frontend ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


# ─── Auth ────────────────────────────────────────────────────────────────────

def _is_hashed(pw: str) -> bool:
    return pw.startswith(("scrypt:", "pbkdf2:", "sha256$"))


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({"ok": False, "error": "Demasiados intentos. Espera un momento."}), 429
    d = request.json or {}
    u = Usuario.query.filter_by(user=d.get("user", ""), activo=True).first()
    if not u:
        return jsonify({"ok": False}), 401

    provided = d.get("pass", "")
    if _is_hashed(u.pass_):
        valid = check_password_hash(u.pass_, provided)
    else:
        # Legacy plaintext — verify then auto-upgrade
        valid = u.pass_ == provided
        if valid:
            u.pass_ = generate_password_hash(provided)
            db.session.commit()

    if not valid:
        return jsonify({"ok": False}), 401

    session.permanent = True
    session["user_id"] = u.id
    session["rol"] = u.rol
    return jsonify({"ok": True, "user": u.user, "nombre": u.nombre, "rol": u.rol})


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"ok": True})


# ─── Máquinas ────────────────────────────────────────────────────────────────

@app.route("/api/maquinas")
@login_required
def get_maquinas():
    return jsonify([m.to_dict(paciente_actual(m.id)) for m in Maquina.query.all()])


@app.route("/api/maquinas", methods=["POST"])
@login_required
def create_maquina():
    d = request.json or {}
    m = Maquina(
        id=next_maq_id(),
        nombre=d["nombre"],
        serial=d.get("serial", ""),
        estado=d.get("estado", "Operativa"),
        ubicacion=d.get("ubicacion", ""),
        notas=d.get("notas", ""),
    )
    db.session.add(m)
    db.session.commit()
    return jsonify(m.to_dict(None)), 201


@app.route("/api/maquinas/<id>", methods=["PUT"])
@login_required
def update_maquina(id):
    m = db.get_or_404(Maquina, id)
    d = request.json or {}
    for f in ("nombre", "serial", "estado", "ubicacion", "notas"):
        if f in d:
            setattr(m, f, d[f])
    db.session.commit()
    return jsonify(m.to_dict(paciente_actual(m.id)))


@app.route("/api/maquinas/<id>", methods=["DELETE"])
@login_required
def delete_maquina(id):
    m = db.get_or_404(Maquina, id)
    for r in Registro.query.filter_by(maquina=id, estatus="Activo").all():
        r.estatus = "Desactivado"
    for r in Registro.query.filter_by(maquina=id).all():
        r.maquina = ""
    db.session.delete(m)
    db.session.commit()
    return jsonify({"ok": True})


# ─── Registros ───────────────────────────────────────────────────────────────

def _registro_from_dict(d, reg_id):
    return Registro(
        id=reg_id,
        fecha=d["fecha"],
        nombre=d["nombre"],
        cedula=d.get("cedula", ""),
        sexo=d.get("sexo", ""),
        edad=d.get("edad"),
        lesion=d.get("lesion", ""),
        direccion=d.get("direccion", ""),
        tel1=d.get("tel1", ""),
        tel2=d.get("tel2", ""),
        dr_refiere=d.get("dr_refiere", ""),
        ars=d.get("ars", "Privado"),
        maquina=d.get("maquina", ""),
        estatus=d.get("estatus", "Activo"),
        motivo=d.get("motivo", ""),
        facturacion=d.get("facturacion", 0),
        saldo_pendiente=d.get("saldo_pendiente", 0),
        lat=d.get("lat"),
        lng=d.get("lng"),
        modo_uso=d.get("modo_uso", ""),
        condicion_salida=d.get("condicion_salida", ""),
        condicion_retorno=d.get("condicion_retorno", ""),
        parametros=d.get("parametros", ""),
        notas=d.get("notas", ""),
        proxima_colocacion=d.get("proxima_colocacion"),
        notas_seguimiento=d.get("notas_seguimiento", ""),
        fecha_retiro=d.get("fecha_retiro"),
        observaciones_retiro=d.get("observaciones_retiro", ""),
        productos=json.dumps(d.get("productos", [])),
        fotos=json.dumps(d.get("fotos", [])),
    )


@app.route("/api/registros")
@login_required
def get_registros():
    regs = Registro.query.order_by(Registro.fecha).all()
    return jsonify([r.to_dict() for r in regs])


@app.route("/api/registros", methods=["POST"])
@login_required
def create_registro():
    d = request.json or {}
    r = _registro_from_dict(d, next_reg_id())
    db.session.add(r)
    db.session.commit()
    return jsonify(r.to_dict()), 201


@app.route("/api/registros/bulk", methods=["POST"])
@login_required
def bulk_create_registros():
    data = request.json or []
    n = next_num(Registro, "R")
    created = []
    for d in data:
        r = _registro_from_dict(d, f"R{n:03d}")
        n += 1
        db.session.add(r)
        created.append(r)
    db.session.commit()
    return jsonify([r.to_dict() for r in created]), 201


@app.route("/api/registros/<id>", methods=["PUT"])
@login_required
def update_registro(id):
    r = db.get_or_404(Registro, id)
    d = request.json or {}
    for f in ("fecha", "nombre", "cedula", "sexo", "edad", "lesion", "direccion",
              "tel1", "tel2", "dr_refiere", "ars", "maquina", "estatus", "motivo",
              "facturacion", "saldo_pendiente", "lat", "lng", "modo_uso",
              "condicion_salida", "condicion_retorno", "parametros", "notas",
              "proxima_colocacion", "notas_seguimiento", "fecha_retiro",
              "observaciones_retiro"):
        if f in d:
            setattr(r, f, d[f])
    if "productos" in d:
        r.productos = json.dumps(d["productos"])
    if "fotos" in d:
        r.fotos = json.dumps(d["fotos"])
    db.session.commit()
    return jsonify(r.to_dict())


@app.route("/api/registros/<id>", methods=["DELETE"])
@login_required
def delete_registro(id):
    r = db.get_or_404(Registro, id)
    db.session.delete(r)
    db.session.commit()
    return jsonify({"ok": True})


# ─── Pacientes (bulk operations) ─────────────────────────────────────────────

@app.route("/api/pacientes/<path:nombre>", methods=["DELETE"])
@login_required
def delete_paciente(nombre):
    Registro.query.filter_by(nombre=nombre).delete()
    pm = PacienteMaster.query.get(nombre)
    if pm:
        db.session.delete(pm)
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/pacientes/<path:nombre>/retiro", methods=["POST"])
@login_required
def retiro_paciente(nombre):
    d = request.json or {}
    activos = Registro.query.filter_by(nombre=nombre, estatus="Activo").all()
    for r in activos:
        r.estatus = "Desactivado"
        r.fecha_retiro = d.get("fecha")
        r.condicion_retorno = d.get("cond", "")
        r.observaciones_retiro = d.get("obs", "")
        if r.maquina:
            m = Maquina.query.get(r.maquina)
            if m:
                cond = d.get("cond", "")
                if cond == "Dañado":
                    m.estado = "Fuera de Servicio"
                elif cond == "Requiere revisión":
                    m.estado = "Requiere Revisión"
    db.session.commit()
    return jsonify({"ok": True})


# ─── Pacientes master ─────────────────────────────────────────────────────────

@app.route("/api/pacientes-master")
@login_required
def get_pacientes_master():
    rows = PacienteMaster.query.all()
    return jsonify({r.nombre: r.to_dict() for r in rows})


@app.route("/api/pacientes-master/<path:nombre>", methods=["PUT"])
@login_required
def update_paciente_master(nombre):
    d = request.json or {}
    nuevo_nombre = d.get("nombre", nombre)

    if nuevo_nombre != nombre:
        for r in Registro.query.filter_by(nombre=nombre).all():
            r.nombre = nuevo_nombre
        pm_old = PacienteMaster.query.get(nombre)
        if pm_old:
            db.session.delete(pm_old)

    pm = PacienteMaster.query.get(nuevo_nombre)
    if not pm:
        pm = PacienteMaster(nombre=nuevo_nombre, datos="{}")
        db.session.add(pm)

    existing = json.loads(pm.datos or "{}")
    existing.update(d)
    pm.datos = json.dumps(existing)
    db.session.commit()
    return jsonify({"ok": True})


# ─── Usuarios ─────────────────────────────────────────────────────────────────

@app.route("/api/usuarios")
@login_required
def get_usuarios():
    # Passwords are never sent to the client
    return jsonify([u.to_dict() for u in Usuario.query.all()])


@app.route("/api/usuarios", methods=["POST"])
@admin_required
def create_usuario():
    d = request.json or {}
    if Usuario.query.filter_by(user=d["user"]).first():
        return jsonify({"error": "Usuario ya existe"}), 409
    u = Usuario(
        id=next_usr_id(),
        user=d["user"],
        pass_=generate_password_hash(d["pass"]),
        nombre=d["nombre"],
        rol=d.get("rol", "usuario"),
        activo=d.get("activo", True),
    )
    db.session.add(u)
    db.session.commit()
    return jsonify(u.to_dict()), 201


@app.route("/api/usuarios/<id>", methods=["PUT"])
@admin_required
def update_usuario(id):
    u = db.get_or_404(Usuario, id)
    d = request.json or {}
    if "nombre" in d:
        u.nombre = d["nombre"]
    if "user" in d:
        dup = Usuario.query.filter_by(user=d["user"]).first()
        if dup and dup.id != id:
            return jsonify({"error": "Usuario ya existe"}), 409
        u.user = d["user"]
    if "pass" in d and d["pass"]:
        u.pass_ = generate_password_hash(d["pass"])
    if "rol" in d:
        u.rol = d["rol"]
    if "activo" in d:
        u.activo = d["activo"]
    db.session.commit()
    return jsonify(u.to_dict())


@app.route("/api/usuarios/<id>", methods=["DELETE"])
@admin_required
def delete_usuario(id):
    u = db.get_or_404(Usuario, id)
    db.session.delete(u)
    db.session.commit()
    return jsonify({"ok": True})


# ─── Import backup JSON ───────────────────────────────────────────────────────

@app.route("/api/importar/backup", methods=["POST"])
@admin_required
def importar_backup():
    d = request.json or {}
    count = {"maquinas": 0, "registros": 0, "usuarios": 0}

    for m_data in d.get("maquinas", []):
        m = Maquina.query.get(m_data["id"])
        if m:
            for f in ("nombre", "serial", "estado", "ubicacion", "notas"):
                m.__setattr__(f, m_data.get(f, getattr(m, f)))
        else:
            db.session.add(Maquina(
                id=m_data["id"],
                nombre=m_data.get("nombre", ""),
                serial=m_data.get("serial", ""),
                estado=m_data.get("estado", "Operativa"),
                ubicacion=m_data.get("ubicacion", ""),
                notas=m_data.get("notas", ""),
            ))
        count["maquinas"] += 1

    for r_data in d.get("registros", []):
        r = Registro.query.get(r_data["id"])
        if r:
            for f in ("fecha","nombre","cedula","sexo","edad","lesion","direccion",
                      "tel1","tel2","dr_refiere","ars","maquina","estatus","motivo",
                      "facturacion","modo_uso","condicion_salida","condicion_retorno",
                      "parametros","notas","proxima_colocacion","notas_seguimiento",
                      "fecha_retiro","observaciones_retiro"):
                if f in r_data:
                    setattr(r, f, r_data[f])
            r.productos = json.dumps(r_data.get("productos", []))
            r.fotos = json.dumps(r_data.get("fotos", []))
        else:
            db.session.add(_registro_from_dict(r_data, r_data["id"]))
        count["registros"] += 1

    for u_data in d.get("usuarios", []):
        u = Usuario.query.get(u_data["id"])
        if u:
            u.nombre = u_data.get("nombre", u.nombre)
            u.user = u_data.get("user", u.user)
            if u_data.get("pass"):
                raw = u_data["pass"]
                u.pass_ = raw if _is_hashed(raw) else generate_password_hash(raw)
            u.rol = u_data.get("rol", u.rol)
            u.activo = u_data.get("activo", u.activo)
        else:
            raw = u_data.get("pass", "")
            db.session.add(Usuario(
                id=u_data["id"],
                user=u_data["user"],
                pass_=raw if _is_hashed(raw) else generate_password_hash(raw),
                nombre=u_data.get("nombre", ""),
                rol=u_data.get("rol", "usuario"),
                activo=u_data.get("activo", True),
            ))
        count["usuarios"] += 1

    for nombre, datos in d.get("pacientes_master", {}).items():
        pm = PacienteMaster.query.get(nombre)
        if pm:
            existing = json.loads(pm.datos or "{}")
            existing.update(datos)
            pm.datos = json.dumps(existing)
        else:
            db.session.add(PacienteMaster(nombre=nombre, datos=json.dumps(datos)))

    db.session.commit()
    return jsonify({"ok": True, "importados": count})


# ─── Export endpoints ─────────────────────────────────────────────────────────

def _all_data():
    regs = [r.to_dict() for r in Registro.query.order_by(Registro.fecha).all()]
    maquinas = [m.to_dict() for m in Maquina.query.all()]
    return regs, maquinas


@app.route("/api/exportar/registros/excel")
@login_required
def exportar_excel():
    regs, maquinas = _all_data()
    buf = generate_excel(regs, maquinas)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dipromes_registros.xlsx"},
    )


@app.route("/api/exportar/registros/csv")
@login_required
def exportar_csv():
    regs, maquinas = _all_data()
    buf = generate_csv(regs, maquinas)
    return Response(
        buf.read(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=dipromes_registros.csv"},
    )


@app.route("/api/exportar/registros/pdf")
@login_required
def exportar_pdf():
    try:
        regs, maquinas = _all_data()
        buf = generate_pdf(regs, maquinas)
    except Exception:
        return jsonify({"error": "PDF no disponible en este plan"}), 503
    return Response(
        buf.read(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=dipromes_registros.pdf"},
    )


@app.route("/api/exportar/maquinas/excel")
@login_required
def exportar_maquinas_excel():
    _, maquinas = _all_data()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Máquinas"
    hf = PatternFill("solid", fgColor="0066CC")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ID", "Nombre", "Serial", "Estado", "Ubicación", "Paciente Actual", "Notas"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.border = brd
        c.alignment = Alignment(horizontal="center")

    for row_idx, m in enumerate(maquinas, 2):
        for col, val in enumerate([
            m["id"], m["nombre"], m["serial"], m["estado"],
            m["ubicacion"], m.get("paciente_actual") or "Disponible", m["notas"],
        ], 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = brd; c.font = Font(size=10)

    for col, w in enumerate([8, 22, 14, 18, 14, 24, 28], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dipromes_maquinas.xlsx"},
    )


# ─── Config (ARS list, etc.) ──────────────────────────────────────────────────

@app.route("/api/config/<key>")
@login_required
def get_config(key):
    c = Config.query.get(key)
    if not c:
        return jsonify(None), 404
    return jsonify(json.loads(c.value))


@app.route("/api/config/<key>", methods=["PUT"])
@admin_required
def set_config(key):
    c = Config.query.get(key)
    payload = json.dumps(request.json, ensure_ascii=False)
    if c:
        c.value = payload
    else:
        db.session.add(Config(key=key, value=payload))
    db.session.commit()
    return jsonify(json.loads(payload))


# ─── DB management ────────────────────────────────────────────────────────────

@app.route("/api/exportar/backup")
@admin_required
def exportar_backup():
    """Full JSON backup — passwords are exported as hashes, never plaintext."""
    regs = [r.to_dict() for r in Registro.query.order_by(Registro.fecha).all()]
    maquinas = [m.to_dict() for m in Maquina.query.all()]
    # include_pass=True is safe here because passwords are now hashed
    usuarios = [u.to_dict(include_pass=True) for u in Usuario.query.all()]
    pacientes_master = {
        pm.nombre: json.loads(pm.datos or "{}")
        for pm in PacienteMaster.query.all()
    }
    payload = json.dumps({
        "version": 2,
        "exportado": __import__("datetime").datetime.now().isoformat(),
        "maquinas": maquinas,
        "registros": regs,
        "usuarios": usuarios,
        "pacientes_master": pacientes_master,
    }, ensure_ascii=False, indent=2)
    return Response(
        payload.encode("utf-8"),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=dipromes_backup.json"},
    )


@app.route("/api/db/reset", methods=["POST"])
@admin_required
def db_reset():
    """Wipe all data. Pass reseed=true to reload demo data."""
    reseed = (request.json or {}).get("reseed", False)
    Registro.query.delete()
    Maquina.query.delete()
    PacienteMaster.query.delete()
    db.session.commit()
    if reseed:
        seed_if_empty()
    return jsonify({"ok": True, "reseed": reseed})


# ─── DB Init & Seed ──────────────────────────────────────────────────────────

SEED_MAQUINAS = [
    {"id": "MAQ01", "nombre": "Maquina No. 1", "serial": "250619003", "estado": "Operativa", "ubicacion": "Consulta", "notas": ""},
    {"id": "MAQ02", "nombre": "Maquina No. 2", "serial": "250619002", "estado": "Operativa", "ubicacion": "Consulta", "notas": ""},
    {"id": "MAQ03", "nombre": "Maquina No. 3", "serial": "250904014", "estado": "Operativa", "ubicacion": "Domicilio", "notas": ""},
    {"id": "MAQ04", "nombre": "Maquina No. 4", "serial": "250317003", "estado": "Operativa", "ubicacion": "Domicilio", "notas": ""},
    {"id": "MAQ05", "nombre": "Maquina No. 5", "serial": "240411004", "estado": "Requiere Revisión", "ubicacion": "Consulta", "notas": "Revisar estado"},
    {"id": "MAQ06", "nombre": "Maquina No. 6", "serial": "250317004", "estado": "Operativa", "ubicacion": "Consulta", "notas": ""},
    {"id": "MAQ07", "nombre": "Maquina No. 7", "serial": "", "estado": "Sin registrar", "ubicacion": "", "notas": "Sin serial"},
    {"id": "MAQ08", "nombre": "Maquina No. 8", "serial": "", "estado": "Sin registrar", "ubicacion": "", "notas": "Sin serial"},
]

SEED_REGISTROS = [
    {"id":"R001","fecha":"2026-05-05","nombre":"Sebastian Adolfo Gomez F.","cedula":"001-0553575-1","sexo":"M","edad":83,"lesion":"Pierna Derecha","direccion":"Hospital de la Policia","tel1":"829-702-2598 (Kelvin Adolfo)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"2da Colocacion","facturacion":0},
    {"id":"R002","fecha":"2026-05-01","nombre":"Dr. Felix Batista","cedula":"","sexo":"","edad":None,"lesion":"Pecho","direccion":"Azilo Haina","tel1":"829-907-1783","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R003","fecha":"2026-05-03","nombre":"Sobeida Mora","cedula":"024-0014327-3","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"(Los Llanos San Pedro)","tel1":"829-469-2280","tel2":"","dr_refiere":"Juan Vicente Mendez","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"2da Colocacion","facturacion":16000},
    {"id":"R004","fecha":"2026-05-05","nombre":"Jose Adames","cedula":"","sexo":"","edad":80,"lesion":"Pie Izquierdo","direccion":"Bella Vista","tel1":"809-330-5529","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"2da colocacion de 5","facturacion":0},
    {"id":"R005","fecha":"2026-05-02","nombre":"Zunilda Cesar Hilson","cedula":"001-0003343-0","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"Ecologica","tel1":"809-268-8181","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R006","fecha":"2026-05-02","nombre":"Arturo Hernandez","cedula":"","sexo":"","edad":None,"lesion":"Zacra","direccion":"Pct Oncologo","tel1":"","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"retirado el viernes 8 de mayo","facturacion":13000},
    {"id":"R007","fecha":"2026-05-06","nombre":"Mercedes Rosario","cedula":"001-0485475-7","sexo":"F","edad":59,"lesion":"Nalgas","direccion":"Hainamosa","tel1":"849-624-7905 (Marcos)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13500},
    {"id":"R008","fecha":"2026-05-08","nombre":"Zunilda Cesar Hilson","cedula":"001-0003343-0","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"Ecologica","tel1":"809-268-8181","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R009","fecha":"2026-05-09","nombre":"Dr. Felix Batista","cedula":"","sexo":"","edad":None,"lesion":"Pecho","direccion":"Azilo Haina","tel1":"829-907-1783","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R010","fecha":"2026-05-10","nombre":"Sobeida Mora","cedula":"024-0014327-3","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"(Los Llanos San Pedro)","tel1":"829-469-2280","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"3era colocacion","facturacion":16000},
    {"id":"R011","fecha":"2026-05-12","nombre":"Jose Adames","cedula":"","sexo":"","edad":80,"lesion":"Pie Izquierdo","direccion":"Bella Vista","tel1":"809-330-5529","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"3era colocacion de 5","facturacion":3000},
    {"id":"R012","fecha":"2026-05-12","nombre":"Sebastian Adolfo Gomez F.","cedula":"001-0553575-1","sexo":"M","edad":83,"lesion":"Pierna Derecha","direccion":"Hospital de la Policia","tel1":"829-702-2598 (Kelvin Adolfo)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"3era Colocacion de 5 (PCT Murio 14 may)","facturacion":0},
    {"id":"R013","fecha":"2026-05-14","nombre":"Mercedes Rosario","cedula":"001-0485475-7","sexo":"F","edad":59,"lesion":"Nalgas","direccion":"Hainamosa","tel1":"849-624-7905 (Marcos)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"retirado el viernes 22 de mayo","facturacion":13000},
    {"id":"R014","fecha":"2026-05-13","nombre":"Ramon Emilio Campusano","cedula":"001-0676784-1","sexo":"","edad":73,"lesion":"Zacra, las 2 caderas","direccion":"Manoguayabo","tel1":"809-304-5187 (Marcos Campusano)","tel2":"809-217-9714 (Victor Campusano)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"1era Colocacion de 5 (Resta RD$35,000.00)","facturacion":40000,"saldo_pendiente":35000},
    {"id":"R015","fecha":"2026-05-15","nombre":"Dr. Felix Batista","cedula":"","sexo":"","edad":None,"lesion":"Pecho","direccion":"Azilo Haina","tel1":"829-907-1783","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R016","fecha":"2026-05-16","nombre":"Sobeida Mora","cedula":"024-0014327-3","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"(Los Llanos San Pedro)","tel1":"829-469-2280","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"4ta Colocacion","facturacion":16000},
    {"id":"R017","fecha":"2026-05-18","nombre":"Faustino Rosario","cedula":"059-0001895-2","sexo":"F","edad":70,"lesion":"Pecho","direccion":"Nagua (Mac Center)","tel1":"809-946-0510 (Angelita Pichardo)","tel2":"849-403-78210 (Paulina)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"1era Colocacion","facturacion":13000},
    {"id":"R018","fecha":"2026-05-19","nombre":"Jose Adames","cedula":"","sexo":"","edad":80,"lesion":"Pie Izquierdo","direccion":"Bella Vista","tel1":"809-330-5529","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"4ta de 5 Colocaciones","facturacion":0},
    {"id":"R019","fecha":"2026-05-21","nombre":"Ramon Emilio Campusano","cedula":"001-0676784-1","sexo":"","edad":73,"lesion":"Zacra, las 2 caderas","direccion":"Manoguayabo","tel1":"809-304-5187 (Marcos Campusano)","tel2":"809-217-9714 (Victor Campusano)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"2da de 5 Colocaciones","facturacion":0},
    {"id":"R020","fecha":"2026-05-21","nombre":"Julia Ozuna Rodriguez","cedula":"001-0449791-2","sexo":"","edad":72,"lesion":"Zacra","direccion":"Res. Maximo Gomez","tel1":"","tel2":"809-568-1314","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R021","fecha":"2026-05-22","nombre":"Dr. Felix Batista","cedula":"","sexo":"","edad":None,"lesion":"Pecho","direccion":"Azilo Haina","tel1":"829-907-1783","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":13000},
    {"id":"R022","fecha":"2026-05-23","nombre":"Julio Mendez Mendez","cedula":"070-0004347-6","sexo":"","edad":67,"lesion":"","direccion":"Barahona","tel1":"829-478-0597 (Miriam)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":22000},
    {"id":"R023","fecha":"2026-05-23","nombre":"Sobeida Mora","cedula":"024-0014327-3","sexo":"F","edad":None,"lesion":"Pierna Izquierda","direccion":"(Los Llanos San Pedro)","tel1":"829-469-2280","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"Retirado el Sabado 30 (Decision Medica)","facturacion":0},
    {"id":"R024","fecha":"2026-05-25","nombre":"Armenio Gomez","cedula":"022-0002732-0","sexo":"","edad":60,"lesion":"Pierna derecha","direccion":"Neiba (Moscoso Puello)","tel1":"829-868-4395 (Lisset)","tel2":"809-663-6823 (Candido)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"","facturacion":12000},
    {"id":"R025","fecha":"2026-05-25","nombre":"Elias German Mago Quezada","cedula":"001-6429938-3","sexo":"M","edad":62,"lesion":"Pierna","direccion":"San Luis (Ney Arias)","tel1":"39 379 159 4946","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"1era Colocacion","facturacion":13000},
    {"id":"R026","fecha":"2026-05-25","nombre":"Faustino Rosario","cedula":"059-0001895-2","sexo":"F","edad":70,"lesion":"Pecho","direccion":"Nagua (Mac Center)","tel1":"809-946-0510 (Angelita Pichardo)","tel2":"849-403-78210 (Paulina)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"2da Colocacion","facturacion":13000},
    {"id":"R027","fecha":"2026-05-27","nombre":"Jose Adames","cedula":"","sexo":"","edad":80,"lesion":"Pie Izquierdo","direccion":"Bella Vista","tel1":"809-330-5529","tel2":"","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"5ta Colocacion (Retirada el 2 de Junios)","facturacion":0},
    {"id":"R028","fecha":"2026-05-27","nombre":"Ramon Emilio Campusano","cedula":"001-0676784-1","sexo":"","edad":73,"lesion":"Zacra, las 2 caderas","direccion":"Manoguayabo","tel1":"809-304-5187 (Marcos Campusano)","tel2":"809-217-9714 (Victor Campusano)","dr_refiere":"","ars":"Privado","maquina":"","estatus":"Desactivado","motivo":"3era Colocacion","facturacion":0},
    {"id":"R029","fecha":"2026-05-29","nombre":"Dr. Felix Batista","cedula":"","sexo":"","edad":None,"lesion":"Pecho","direccion":"Azilo Haina","tel1":"829-907-1783","tel2":"","dr_refiere":"","ars":"Privado","maquina":"MAQ01","estatus":"Activo","motivo":"","facturacion":13000},
    {"id":"R030","fecha":"2026-05-30","nombre":"Santo Tejeda","cedula":"003-0065067-8","sexo":"","edad":52,"lesion":"Pierna derecha","direccion":"Bani","tel1":"829-380-1747 (Orquidea)","tel2":"829-788-2378 (Robert)","dr_refiere":"","ars":"Privado","maquina":"MAQ04","estatus":"Activo","motivo":"1era Colocacion","facturacion":13000},
    {"id":"R031","fecha":"2026-05-30","nombre":"Julio Mendez Mendez","cedula":"070-0004347-6","sexo":"","edad":67,"lesion":"","direccion":"Barahona","tel1":"829-478-0597 (Miriam)","tel2":"","dr_refiere":"","ars":"Privado","maquina":"MAQ03","estatus":"Activo","motivo":"2da Colocacacion","facturacion":22000},
    {"id":"R032","fecha":"2026-05-30","nombre":"Julia Ozuna Rodriguez","cedula":"001-0449791-2","sexo":"","edad":72,"lesion":"Zacra","direccion":"Res. Maximo Gomez","tel1":"","tel2":"809-568-1314","dr_refiere":"","ars":"Privado","maquina":"MAQ02","estatus":"Activo","motivo":"2da Colocacion","facturacion":13000},
]

# Seed passwords come from env vars — change via Render dashboard, not in code
_ADMIN_PASS = os.environ.get("ADMIN_PASS", "dipromes2026")
_DR1_PASS = os.environ.get("DR1_PASS", "doctor123")

SEED_ARS = [
    "Senasa Subsidiado", "Senasa Contributivo", "Semma",
    "ARS Humano", "ARS Universal", "ARS Simag", "ARS Reservas",
    "ARS Meta Salud", "ARS Mapfre BHD", "ARS CMD", "ARS Futuro",
    "ARS Plan Salud Banco Central", "ARS APS", "ARS UASD",
    "ARS Premier", "ARS LAR", "ARS Yunen", "ARS Asemap",
    "ARS Monumental", "Privado (Sin ARS)",
]


def seed_if_empty():
    if Maquina.query.count() == 0:
        for m in SEED_MAQUINAS:
            db.session.add(Maquina(**m))
    if Registro.query.count() == 0:
        for r in SEED_REGISTROS:
            db.session.add(_registro_from_dict(r, r["id"]))
    if Usuario.query.count() == 0:
        db.session.add(Usuario(
            id="U001", user="admin",
            pass_=generate_password_hash(_ADMIN_PASS),
            nombre="Administrador", rol="admin", activo=True,
        ))
        db.session.add(Usuario(
            id="U002", user="dr1",
            pass_=generate_password_hash(_DR1_PASS),
            nombre="Dr. Médico", rol="usuario", activo=True,
        ))
    if not Config.query.get("ars"):
        db.session.add(Config(key="ars", value=json.dumps(SEED_ARS, ensure_ascii=False)))
    db.session.commit()


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        seed_if_empty()
    app.run(host="127.0.0.1", port=5000, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
