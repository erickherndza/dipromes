"""Export helpers: Excel, CSV, PDF, Consent HTML."""
import io
import csv
from datetime import datetime


def fmt_dop(v):
    return f"RD${v:,.0f}" if v else "—"


def fmt_fecha(d):
    if not d:
        return "—"
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return d


# ── Excel ──────────────────────────────────────────────────────────────────
def generate_excel(registros, maquinas):
    """Return BytesIO with an .xlsx workbook containing registros and machines."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Registro Pacientes ──
    ws = wb.active
    ws.title = "Registro Pacientes"

    header_fill = PatternFill("solid", fgColor="0066CC")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Fecha", "Nombre", "Cédula", "Sexo", "Edad",
        "Área de Lesión", "Dirección", "Teléfono 1", "Teléfono 2",
        "Dr. que refiere", "ARS", "Máquina", "Estatus",
        "Colocación / Motivo", "Conduce (DOP)", "Fecha Retiro",
        "Condición Retorno", "Modo Uso", "Observaciones",
    ]
    maq_map = {m['id']: m['nombre'] for m in maquinas}

    ws.row_dimensions[1].height = 30
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, r in enumerate(registros, 2):
        row_data = [
            r.get("fecha", ""),
            r.get("nombre", ""),
            r.get("cedula", ""),
            r.get("sexo", ""),
            r.get("edad", ""),
            r.get("lesion", ""),
            r.get("direccion", ""),
            r.get("tel1", ""),
            r.get("tel2", ""),
            r.get("dr_refiere", ""),
            r.get("ars", ""),
            maq_map.get(r.get("maquina", ""), r.get("maquina", "")),
            r.get("estatus", ""),
            r.get("motivo", ""),
            r.get("facturacion", 0),
            r.get("fecha_retiro", ""),
            r.get("condicion_retorno", ""),
            r.get("modo_uso", ""),
            r.get("observaciones_retiro", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.font = Font(size=10)
            if r.get("estatus") == "Activo":
                cell.fill = PatternFill("solid", fgColor="E6F5EF")

    # Column widths
    col_widths = [12, 28, 16, 6, 6, 20, 28, 18, 18, 20, 12, 14, 12, 24, 14, 12, 18, 14, 24]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Inventario de Máquinas ──
    ws2 = wb.create_sheet("Inventario de Maquinas")
    maq_headers = ["ID", "Nombre", "Serial", "Estado", "Ubicación", "Notas"]
    ws2.row_dimensions[1].height = 25
    for col, h in enumerate(maq_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, m in enumerate(maquinas, 2):
        for col, val in enumerate([
            m.get("id"), m.get("nombre"), m.get("serial"),
            m.get("estado"), m.get("ubicacion"), m.get("notas"),
        ], 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.font = Font(size=10)

    for col, w in enumerate([10, 22, 16, 18, 14, 28], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── CSV ─────────────────────────────────────────────────────────────────────
def generate_csv(registros, maquinas):
    """Return BytesIO with UTF-8 BOM CSV (Excel-compatible)."""
    maq_map = {m['id']: m['nombre'] for m in maquinas}
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Fecha", "Nombre", "Cédula", "Sexo", "Edad",
        "Área de Lesión", "Dirección", "Teléfono 1", "Teléfono 2",
        "Dr. que refiere", "ARS", "Máquina", "Estatus",
        "Colocación / Motivo", "Conduce (DOP)", "Fecha Retiro",
        "Condición Retorno", "Modo Uso", "Observaciones",
    ])
    for r in registros:
        writer.writerow([
            r.get("fecha", ""),
            r.get("nombre", ""),
            r.get("cedula", ""),
            r.get("sexo", ""),
            r.get("edad", "") or "",
            r.get("lesion", ""),
            r.get("direccion", ""),
            r.get("tel1", ""),
            r.get("tel2", ""),
            r.get("dr_refiere", ""),
            r.get("ars", ""),
            maq_map.get(r.get("maquina", ""), r.get("maquina", "")),
            r.get("estatus", ""),
            r.get("motivo", ""),
            r.get("facturacion", 0),
            r.get("fecha_retiro", "") or "",
            r.get("condicion_retorno", ""),
            r.get("modo_uso", ""),
            r.get("observaciones_retiro", ""),
        ])
    # UTF-8 BOM so Excel opens it correctly
    raw = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return io.BytesIO(raw)


# ── PDF ─────────────────────────────────────────────────────────────────────
def generate_pdf(registros, maquinas, titulo="Registro de Pacientes"):
    """Return BytesIO with a PDF report."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    brand = colors.HexColor("#0066CC")
    brand_dark = colors.HexColor("#004499")
    ok_color = colors.HexColor("#1a9b6c")
    warn_color = colors.HexColor("#d4860a")

    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontSize=16, textColor=brand_dark, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#4a5568"), spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        "cell", parent=styles["Normal"], fontSize=7.5, leading=10,
    )

    maq_map = {m['id']: m['nombre'] for m in maquinas}
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    story = [
        Paragraph("DIPROMES", title_style),
        Paragraph(f"{titulo} — Generado el {now}", sub_style),
        HRFlowable(width="100%", thickness=1, color=brand, spaceAfter=10),
    ]

    # Summary metrics
    activos = sum(1 for r in registros if r.get("estatus") == "Activo")
    total_fact = sum(r.get("facturacion", 0) or 0 for r in registros)
    metrics_data = [[
        Paragraph(f"<b>Total registros</b><br/>{len(registros)}", cell_style),
        Paragraph(f"<b>Activos</b><br/>{activos}", cell_style),
        Paragraph(f"<b>Total conduces</b><br/>RD${total_fact:,.0f}", cell_style),
        Paragraph(f"<b>Máquinas</b><br/>{len(maquinas)}", cell_style),
    ]]
    metrics_table = Table(metrics_data, colWidths=[5 * cm] * 4)
    metrics_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f4f6f8")),
        ("ROUNDEDCORNERS", [4]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d0d7de")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 0.4 * cm))

    # Main table
    col_headers = [
        "Fecha", "Paciente", "Cédula", "Lesión",
        "Máquina", "Colocación", "Estatus", "Conduce",
    ]
    header_row = [Paragraph(f"<b>{h}</b>", cell_style) for h in col_headers]
    table_data = [header_row]

    for r in registros:
        maq_label = maq_map.get(r.get("maquina", ""), r.get("maquina", "") or "—")
        fact = r.get("facturacion", 0) or 0
        estatus = r.get("estatus", "")
        table_data.append([
            Paragraph(fmt_fecha(r.get("fecha")), cell_style),
            Paragraph(r.get("nombre", ""), cell_style),
            Paragraph(r.get("cedula", "") or "—", cell_style),
            Paragraph(r.get("lesion", "") or "—", cell_style),
            Paragraph(maq_label or "—", cell_style),
            Paragraph(r.get("motivo", "") or "—", cell_style),
            Paragraph(estatus, cell_style),
            Paragraph(f"RD${fact:,.0f}" if fact else "—", cell_style),
        ])

    col_widths = [2.2 * cm, 5.5 * cm, 3 * cm, 3.5 * cm, 3 * cm, 4 * cm, 2.3 * cm, 2.5 * cm]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    ts = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        # Body
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d0d7de")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])

    # Highlight active rows
    for i, r in enumerate(registros, 1):
        if r.get("estatus") == "Activo":
            ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#e6f5ef"))
            ts.add("TEXTCOLOR", (6, i), (6, i), ok_color)

    tbl.setStyle(ts)
    story.append(tbl)

    # Footer note
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"<font color='#9aa5b4'>DIPROMES · Documento generado automáticamente · {now}</font>",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, alignment=1),
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ── Consent HTML (printable, no extra deps) ──────────────────────────────────
def consent_html(c: dict) -> str:
    """Return a self-contained printable HTML page for a consent record."""
    def _line(val: str, width: str = "260px") -> str:
        content = val if val else "&nbsp;" * 30
        return (
            f'<span style="display:inline-block;min-width:{width};'
            f'border-bottom:1px solid #333;padding-bottom:1px;'
            f'font-weight:{\"600\" if val else \"400\"}">{content}</span>'
        )

    fecha = c.get("fecha_firma", "")
    dd, mm, yyyy = ("", "", "")
    if fecha and len(fecha) == 10:
        yyyy, mm, dd = fecha.split("-")

    edad_str = f"{c['edad']} años" if c.get("edad") else ""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Consentimiento Informado — {c.get('nombre','')}</title>
<style>
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:'Times New Roman',serif;font-size:11pt;color:#1a1a1a;background:#fff;padding:18mm 20mm}}
  .header{{text-align:center;margin-bottom:14px;border-bottom:2px solid #7B1A1A;padding-bottom:10px}}
  .logo{{font-size:18pt;font-weight:700;color:#7B1A1A;letter-spacing:1px}}
  .info{{font-size:9pt;color:#444;margin-top:3px;line-height:1.5}}
  h1{{text-align:center;font-size:13pt;color:#7B1A1A;text-transform:uppercase;letter-spacing:1px;margin:14px 0 4px}}
  h2{{font-size:10pt;color:#7B1A1A;text-transform:uppercase;letter-spacing:.5px;border-left:3px solid #7B1A1A;padding-left:7px;margin:14px 0 7px}}
  .row{{margin-bottom:6px;font-size:10.5pt;line-height:1.8}}
  .lbl{{font-weight:600;margin-right:4px}}
  ul{{padding-left:20px;margin:4px 0}}
  li{{margin-bottom:3px;font-size:10.5pt}}
  p{{font-size:10.5pt;margin-bottom:4px;line-height:1.6}}
  .firma-row{{display:flex;gap:30px;margin-top:8px}}
  .firma-box{{flex:1}}
  .firma-line{{border-bottom:1px solid #333;margin-bottom:4px;height:28px}}
  .firma-lbl{{font-size:9.5pt;color:#444}}
  .footer{{margin-top:16px;border-top:1px solid #ccc;padding-top:6px;text-align:center;font-size:8.5pt;color:#777}}
  @media print{{body{{padding:10mm 14mm}}@page{{size:A4 portrait;margin:10mm 14mm}}}}
</style>
</head>
<body>
<div class="header">
  <div class="logo">DIPROMES</div>
  <div class="info">
    Calle 6 Santo Tomás de Aquino No. 55 · Zona Universitaria, Santo Domingo, Rep. Dom.<br>
    RNC: 131950965 &nbsp;|&nbsp; 829-715-6235 · 829-630-1776 · 829-902-3716 · 809-706-4792
  </div>
</div>

<h1>Consentimiento Informado<br><span style="font-size:10pt">Terapia VAC</span></h1>

<h2>Datos del Paciente</h2>
<div class="row"><span class="lbl">Nombre:</span>{_line(c.get('nombre',''),'280px')} &nbsp; <span class="lbl">Cédula:</span>{_line(c.get('cedula',''),'160px')}</div>
<div class="row"><span class="lbl">Edad:</span>{_line(edad_str,'70px')} &nbsp; <span class="lbl">Dirección:</span>{_line(c.get('direccion',''),'220px')} &nbsp; <span class="lbl">Tel.:</span>{_line(c.get('telefono',''),'130px')}</div>

<h2>Indicación Médica</h2>
<div class="row"><span class="lbl">Nombre del médico:</span>{_line(c.get('medico',''),'200px')} &nbsp; <span class="lbl">Centro de salud:</span>{_line(c.get('centro_salud',''),'185px')}</div>

<h2>Información del Procedimiento</h2>
<p>He recibido una explicación clara sobre la <strong>Terapia VAC</strong>. Se me ha explicado y entiendo en qué consiste el tratamiento con el objetivo de:</p>
<ul>
  <li>Ayudar a la cicatrización</li>
  <li>Eliminar secreciones o líquidos acumulados</li>
  <li>Disminuir el riesgo de infección</li>
  <li>Favorecer la formación de tejido sano</li>
</ul>
<p style="margin-top:6px">También comprendo que este procedimiento será realizado por personal capacitado de <strong>Dipromes Terapias VAC</strong>, siguiendo las indicaciones médicas correspondientes y que los beneficios esperados van a depender de la fisiopatología del paciente y su lesión.</p>

<h2>Posibles Riesgos y Molestias</h2>
<p>Se me ha explicado que pueden presentarse algunas molestias o situaciones durante el tratamiento, tales como:</p>
<ul>
  <li>Dolor o incomodidad en la zona tratada</li>
  <li>Irritación de la piel</li>
  <li>Sangrado leve</li>
  <li>Necesidad de ajustes o cambios en el tratamiento</li>
</ul>

<h2>Mis Responsabilidades como Paciente</h2>
<p>Me comprometo a:</p>
<ul>
  <li>Seguir las indicaciones del personal de salud</li>
  <li>No manipular el equipo sin autorización</li>
  <li>Informar cualquier molestia, dolor o cambio en mi herida</li>
  <li>Asistir a las citas y seguimientos indicados</li>
</ul>

<h2>Uso de Información y Datos</h2>
<p>Autorizo a <strong>Dipromes Terapias VAC</strong> a utilizar mis datos personales y clínicos únicamente para fines médicos, administrativos y de seguimiento de mi tratamiento.</p>

<h2>Declaración de Consentimiento</h2>
<p>Declaro que:</p>
<ul>
  <li>He recibido información clara, suficiente y comprensible sobre el procedimiento.</li>
  <li>Estoy de acuerdo con lo que se ha explicado y estoy consciente de lo que significa el tratamiento.</li>
  <li>Entiendo los beneficios, riesgos y posibles complicaciones.</li>
</ul>

<h2>Firmas</h2>
<div class="firma-row">
  <div class="firma-box">
    <div class="firma-line"></div>
    <div class="firma-lbl">Firma del paciente o familiar autorizado</div>
    <div style="margin-top:6px;font-size:10pt"><span class="lbl">Fecha:</span> {dd} / {mm} / {yyyy}</div>
  </div>
  <div class="firma-box">
    <div class="firma-line"></div>
    <div class="firma-lbl">Personal de salud (nombre y firma)</div>
  </div>
</div>

<div class="footer">DIPROMES Terapias VAC · Calle 6 Santo Tomás de Aquino No. 55, Zona Universitaria, Santo Domingo · RNC 131950965</div>
</body>
</html>"""
